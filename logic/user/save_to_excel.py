import pandas as pd
from datetime import datetime
import os

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter

RESULTS_FILE = "survey_results/all_results.xlsx"




async def save_to_excel(all_answers: dict, username: str, poll_name: str) -> None:
    os.makedirs('survey_results', exist_ok=True)

    user_data = {
        "юзернейм": username,
        "Опрос": poll_name,
        "Дата и время": datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # Добавляем ответы типа 1_to_5, если они есть
    if "1_to_5_answers" in all_answers and all_answers["1_to_5_answers"]:
        for question, answer in all_answers["1_to_5_answers"].items():
            user_data[f"[1-5] {question}"] = answer

    # Добавляем открытые ответы, если они есть
    if "open_answers" in all_answers and all_answers["open_answers"]:
        for question, answer in all_answers["open_answers"].items():
            user_data[f"[текст] {question}"] = answer

    new_data = pd.DataFrame([user_data])

    if os.path.exists(RESULTS_FILE):
        existing_data = pd.read_excel(RESULTS_FILE, engine='openpyxl')

        combined_data = pd.concat([existing_data, new_data], ignore_index=True)
    else:
        combined_data = new_data

    combined_data.to_excel(RESULTS_FILE, index=False, engine='openpyxl')

    adjust_column_width(RESULTS_FILE)


def get_results():
    if not os.path.exists(RESULTS_FILE):
        open(RESULTS_FILE, "w").close()
    if not os.path.getsize(RESULTS_FILE) > 0:
        return []
    data = pd.read_excel(RESULTS_FILE,engine='openpyxl')
    data["Дата и время"] = pd.to_datetime(data["Дата и время"])
    cur_mon = datetime.now().month
    data = data[data["Дата и время"].dt.month == cur_mon]
    return data.to_dict(orient='records')


def adjust_column_width(file_path):
    """Функция для автоматической подгонки ширины столбцов"""
    wb = load_workbook(file_path)
    ws = wb.active

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        # Находим максимальную длину содержимого в столбце
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Устанавливаем ширину с небольшим запасом
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
